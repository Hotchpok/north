import pandas as pd
from module3 import get_data_by_surname
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import numbers

def apply_styles(worksheet):
    # Настройка стилей для заголовков
    header_font = Font(name='Calibri', size=11, bold=True, color='000000')  # Черный текст
    header_alignment = Alignment(horizontal='center', vertical='center')  # Выравнивание по центру
    header_border = Border(
        left=Side(style='thin', color='000000'),  # Тонкая черная граница
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Цвета для заголовков
    date_header_fill = PatternFill(start_color='BDD6EE', end_color='BDD6EE', fill_type='solid')  # Голубой (A1)
    yellow_header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Желтый (тц - Высота кузова)

    # Применяем стиль к заголовкам
    for col in worksheet.iter_cols(1, worksheet.max_column):
        for cell in col:
            if cell.row == 1:  # Заголовки находятся в первой строке
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = header_border
                if cell.column == 1:  # Ячейка A1 (Дата)
                    cell.fill = date_header_fill
                else:  # Остальные заголовки (тц - Высота кузова)
                    cell.fill = yellow_header_fill

    # Настройка стилей для данных
    data_font = Font(name='Calibri', size=11, color='000000')  # Черный текст
    data_alignment = Alignment(horizontal='left', vertical='center')  # Выравнивание по левому краю
    data_border = Border(
        left=Side(style='thin', color='000000'),  # Тонкая черная граница
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    # Цвета для данных
    date_column_fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')  # Зеленый (столбец с датой)
    green_data_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')  # Светло-зеленый (данные под тц, Компания, Время заказа, Тоннаж, особенности)

    # Применяем стиль к данным
    for row in worksheet.iter_rows(2, worksheet.max_row):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = data_border
            if cell.column == 1:  # Столбец с датой (A)
                cell.fill = date_column_fill
                # Устанавливаем формат ячейки на "Дата" (без времени)
                cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2
            elif cell.column == 4:  # Столбец "Время заказа" (D)
                cell.fill = green_data_fill
                # Устанавливаем формат ячейки на "Время" (часы:минуты)
                cell.number_format = 'hh:mm'
            elif cell.column in [2, 3, 5, 6]:  # Столбцы тц (B), Компания (C), Тоннаж (E), особенности (F)
                cell.fill = green_data_fill

    # Настройка ширины столбцов
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter  # Получаем букву столбца
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width

def process_excel(file_path):
    try:
        # Чтение Excel-файла
        if file_path.endswith('.xls'):
            # Используем xlrd для чтения .xls файлов
            df = pd.read_excel(file_path, engine='xlrd')
        else:
            # Используем openpyxl для чтения .xlsx файлов
            df = pd.read_excel(file_path, engine='openpyxl')

        # Проверяем наличие столбца 'особенности'
        if 'особенности' not in df.columns:
            raise ValueError("Столбец 'особенности' не найден в файле.")

        # Убираем строки с пустыми значениями в столбце 'особенности'
        df = df.dropna(subset=['особенности'])

        # Применяем функцию get_data_by_surname к каждой фамилии в столбце 'особенности'
        df['Данные'] = df['особенности'].apply(get_data_by_surname)

        # Разделяем данные на отдельные столбцы
        df['Ф.И.О. Водителя'] = df['Данные'].apply(lambda x: x.get('ФИО', ''))
        df['Номер ТС'] = df['Данные'].apply(lambda x: x.get('Номер ТС', ''))
        df['Телефон водителя'] = df['Данные'].apply(lambda x: x.get('Телефон', ''))
        df['Паспортные данные водителя'] = df['Данные'].apply(lambda x: x.get('Паспортные данные', ''))
        df['Паллетность'] = df['Данные'].apply(lambda x: x.get('Паллетность', ''))
        df['Марка ТС'] = df['Данные'].apply(lambda x: x.get('Марка ТС', ''))
        df['В/У'] = df['Данные'].apply(lambda x: x.get('В/У', ''))

        # Удаляем временный столбец 'Данные'
        df.drop(columns=['Данные'], inplace=True)

        # Оставляем только нужные столбцы
        required_columns = [
            'Дата', 'тц', 'Компания', 'Время заказа', 'Тоннаж', 'особенности',
            'Ф.И.О. Водителя', 'Номер ТС', 'Телефон водителя', 'Паспортные данные водителя',
            'Паллетность', 'Марка ТС', 'В/У', 'Высота кузова'
        ]
        df = df[required_columns]

        # Сохранение изменений в новый файл
        output_file = "output.xlsx"
        df.to_excel(output_file, index=False, engine='openpyxl')

        # Применяем стили к выходному файлу
        from openpyxl import load_workbook
        workbook = load_workbook(output_file)
        worksheet = workbook.active
        apply_styles(worksheet)
        workbook.save(output_file)

        return output_file

    except Exception as e:
        print(f"Ошибка при обработке файла: {e}")
        return None